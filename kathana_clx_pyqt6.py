import os
import shutil
import openpyxl
import stat
import logging
import asyncio
import aiofiles
from openpyxl import Workbook
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import timedelta

logging.basicConfig(level=logging.DEBUG, format='%(message)s')
logger = logging.getLogger()

KATHANA_VERSIONS = [r"B:\\Kathana\\Kathana-Global", r"B:\\Kathana\\Kathana2", r"B:\\Kathana\\Kathana3",
                    r"B\\Kathana\\Kathana3.2", r"B\\Kathana\\Kathana4", r"B\\Kathana\\Kathana5.2",
                    r"B\\Kathana\\Kathana6"]

LOG_XLSX_FILENAME = "KATHANA_LOGS.xlsx"
LOG_XLSX_PATH = os.path.join(os.getcwd(), LOG_XLSX_FILENAME)

def initialize_log_workbook():
    wb_log = Workbook()
    error_log_ws = wb_log.create_sheet('ERROR_LOGS')
    success_log_ws = wb_log.create_sheet('SUCCESS_LOGS')
    default_sheet = wb_log.active
    wb_log.remove(default_sheet)
    wb_log.save(LOG_XLSX_PATH)
    return wb_log

def ensure_directory_exists(path):
    if not os.path.exists(path):
        os.makedirs(path)

ensure_directory_exists(os.path.dirname(LOG_XLSX_PATH))

wb_log = initialize_log_workbook()
error_log_ws = wb_log['ERROR_LOGS']
success_log_ws = wb_log['SUCCESS_LOGS']

def log_error(message):
    logger.error(message)
    error_log_ws.append([message])
    wb_log.save(LOG_XLSX_PATH)

def log_success(message):
    logger.info(message)
    success_log_ws.append([message])
    wb_log.save(LOG_XLSX_PATH)

async def copy_file_async(src_file, dest_file, semaphore):
    async with semaphore:
        if os.path.isfile(src_file):
            try:
                async with aiofiles.open(src_file, 'rb') as src, aiofiles.open(dest_file, 'wb') as dest:
                    await dest.write(await src.read())
                os.chmod(dest_file, stat.S_IWRITE)
                log_success(f"Copied {src_file} to {dest_file}")
            except Exception as e:
                log_error(f"Error copying {src_file} to {dest_file}: {e}")
        else:
            log_error(f"File not found: {src_file}")

async def copy_entity_files(workbook, version_path, entity_type):
    logger.debug(f"Entering copy_entity_files function with version_path: {version_path} and entity_type: {entity_type}")
    sheet_name = entity_type
    version_name = os.path.basename(version_path)

    if sheet_name not in workbook.sheetnames:
        log_error(f"Sheet {sheet_name} not found in the workbook.")
        return

    ws = workbook[sheet_name]
    semaphore = asyncio.Semaphore(50)
    tasks = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        entity_id = row[0]
        folder_name = row[1]
        if not folder_name:
            log_error(f"Missing Folder_Name in row: {row}")
            continue

        dest_dir = os.path.join(r"B:\\Kathana-Out\\Sorted", version_name, entity_type, folder_name)
        ensure_directory_exists(dest_dir)
        files_copied = False

        # Copy Mesh files
        for mesh_file in row[2:6]:  # Mesh1 to Mesh4
            if mesh_file:
                src_file = os.path.join(version_path, "resource", "object", entity_type, "Mesh", mesh_file)
                dest_file = os.path.join(dest_dir, mesh_file)
                tasks.append(copy_file_async(src_file, dest_file, semaphore))
                files_copied = True

        # Copy Animation files
        for ani_file in row[6:]:  # Ani1 to Ani70
            if ani_file:
                src_file = os.path.join(version_path, "resource", "object", entity_type, "Ani", ani_file)
                dest_file = os.path.join(dest_dir, ani_file)
                tasks.append(copy_file_async(src_file, dest_file, semaphore))
                files_copied = True

        if not files_copied:
            shutil.rmtree(dest_dir)
            log_error(f"Removed empty directory: {dest_dir}")

    await asyncio.gather(*tasks)

def copy_and_sort_files(version_path, entity_type):
    logger.debug(f"Entering copy_and_sort_files function with version_path: {version_path} and entity_type: {entity_type}")
    logger.info(f"Copying and sorting {entity_type} files from {version_path}...")
    wb = openpyxl.load_workbook(ENTITY_XLSX_PATH)

    start_time = time.time()

    with ThreadPoolExecutor(max_workers=100) as executor:
        future = executor.submit(asyncio.run, copy_entity_files(wb, version_path, entity_type))
        future.result()

    end_time = time.time()
    elapsed_time = end_time - start_time
    logger.info(f"{entity_type} files copied and sorted. Time elapsed: {str(timedelta(seconds=elapsed_time))}")

def generate_fbx_files(version_path, entity_type, generate_batch_only=False, combined_batch=False, batch_commands=[]):
    logger.debug(f"Entering generate_fbx_files function with version_path: {version_path} and entity_type: {entity_type}")
    logger.info(f"Generating {entity_type} FBX files from {version_path}...")

    root_dir = os.path.join(r"B:\\Kathana-Out\\Sorted", os.path.basename(version_path), entity_type)
    fbx_base_dir = os.path.join(r"B:\\Kathana-Out\\FBX", os.path.basename(version_path), entity_type)
    ensure_directory_exists(fbx_base_dir)

    if combined_batch:
        for root, dirs, files in os.walk(root_dir):
            for file in files:
                if file.endswith(".tmb"):
                    tmb_path = os.path.join(root, file)
                    tab_files = [f for f in files if f.endswith(".tab")]
                    for tab_file in tab_files:
                        tab_path = os.path.join(root, tab_file)
                        output_file = os.path.join(fbx_base_dir, os.path.relpath(tab_path, root_dir)).replace(
                            ".tab", ".fbx"
                        )
                        ensure_directory_exists(os.path.dirname(output_file))
                        command = f'"{NOESIS_EXE_PATH}" ?cmode "{tmb_path}" "{output_file}" -loadanimsingle "{tab_path}" -export -bakeanimscale -showstats -animbonenamematch -fbxnoextraframe'
                        batch_commands.append(command)
    else:
        batch_file_path = os.path.join(root_dir, f"generate_{entity_type.lower()}_fbx.bat")

        with open(batch_file_path, 'w') as batch_file:
            for root, dirs, files in os.walk(root_dir):
                for file in files:
                    if file.endswith(".tmb"):
                        tmb_path = os.path.join(root, file)
                        tab_files = [f for f in files if f.endswith(".tab")]
                        for tab_file in tab_files:
                            tab_path = os.path.join(root, tab_file)
                            output_file = os.path.join(fbx_base_dir, os.path.relpath(tab_path, root_dir)).replace(
                                ".tab", ".fbx"
                            )
                            ensure_directory_exists(os.path.dirname(output_file))
                            command = f'"{NOESIS_EXE_PATH}" ?cmode "{tmb_path}" "{output_file}" -loadanimsingle "{tab_path}" -export -bakeanimscale -showstats -animbonenamematch -fbxnoextraframe'
                            batch_file.write(command + '\n')

        logger.info(f"Batch script for generating {entity_type} FBX files created at {batch_file_path}")
        if not generate_batch_only:
            os.system(f'cmd /c "{batch_file_path}"')
            logger.info(f"{entity_type} FBX files generation complete.")

def generate_combined_fbx_batch_file(version_path):
    logger.debug(f"Entering generate_combined_fbx_batch_file function with version_path: {version_path}")
    batch_commands = []
    generate_fbx_files(
        version_path, 'PC', generate_batch_only=True, combined_batch=True, batch_commands=batch_commands
    )
    generate_fbx_files(
        version_path, 'NPC', generate_batch_only=True, combined_batch=True, batch_commands=batch_commands
    )
    generate_fbx_files(
        version_path, 'Monster', generate_batch_only=True, combined_batch=True, batch_commands=batch_commands
    )

    combined_batch_file_path = os.path.join(
        r"B:\\Kathana-Out\\Sorted", os.path.basename(version_path), "generate_all_fbx.bat"
    )

    with open(combined_batch_file_path, 'w') as batch_file:
        for command in batch_commands:
            batch_file.write(command + '\n')

    logger.info(f"Combined batch script for generating all entity FBX files created at {combined_batch_file_path}")

def clean_up():
    logger.debug("Entering clean_up function")
    sorted_path = r"B:\\Kathana-Out\\Sorted"
    fbx_path = r"B:\\Kathana-Out\\FBX"
    if os.path.exists(sorted_path):
        shutil.rmtree(sorted_path)
        logger.info("Cleaned up the kathana-res-sorted folder")
    else:
        logger.info("kathana-res-sorted folder does not exist")
    if os.path.exists(fbx_path):
        shutil.rmtree(fbx_path)
        logger.info("Cleaned up the kathana-res-fbx folder")
    else:
        logger.info("kathana-res-fbx folder does not exist")
