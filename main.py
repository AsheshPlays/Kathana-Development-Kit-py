import sys
import os
import shutil
import time

import openpyxl
import stat
import asyncio
import aiofiles
from PyQt6.QtWidgets import QApplication, QWidget, QVBoxLayout, QHBoxLayout, QPushButton, QLabel, QTabWidget, QMessageBox, QSizePolicy, QProgressBar
from PyQt6.QtGui import QPixmap, QIcon, QFont
from PyQt6.QtCore import Qt, QThread, pyqtSignal
from openpyxl import Workbook
import logging
from datetime import timedelta

# Initialize logging
logging.basicConfig(level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger()

# Define the base path to resource files
def resource_path(relative_path):
    """ Get the absolute path to the resource, works for dev and for PyInstaller """
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

# Load resources
icon_path = resource_path('asheshicon.png')
banner_path = resource_path('asheshdevkitbanner.png')

# Project Constants
KATHANA_DISPLAY_NAMES = ["Kathana Global", "Kathana 2", "Kathana 3", "Kathana 3.2", "Kathana 4", "Kathana 5.2", "Kathana 6"]
KATHANA_VERSIONS = [r"B:\\Kathana\\Kathana-Global", r"B:\\Kathana\\Kathana2", r"B:\\Kathana\\Kathana3", r"B:\\Kathana\\Kathana3.2", r"B:\\Kathana\\Kathana4", r"B:\\Kathana\\Kathana5.2", r"B:\\Kathana\\Kathana6"]

LOG_XLSX_FILENAME = "KATHANA_LOGS.xlsx"
LOG_XLSX_PATH = os.path.join(os.getcwd(), LOG_XLSX_FILENAME)
ENTITY_XLSX_PATH = r"B:\\Kathana\\Kathana_Entity_PS.xlsx"
NOESIS_EXE_PATH = r"B:\\Kathana\\_Noesis\\Noesis.exe"

def initialize_log_workbook():
    """Initialize the log workbook with sheets for error and success logs."""
    wb_log = Workbook()
    error_log_ws = wb_log.create_sheet('ERROR_LOGS')
    success_log_ws = wb_log.create_sheet('SUCCESS_LOGS')
    default_sheet = wb_log.active
    wb_log.remove(default_sheet)
    wb_log.save(LOG_XLSX_PATH)
    return wb_log

def ensure_directory_exists(path):
    """Ensure that the specified directory exists, creating it if necessary."""
    if not os.path.exists(path):
        os.makedirs(path)
        logger.debug(f"Created directory: {path}")

ensure_directory_exists(os.path.dirname(LOG_XLSX_PATH))

wb_log = initialize_log_workbook()
error_log_ws = wb_log['ERROR_LOGS']
success_log_ws = wb_log['SUCCESS_LOGS']

def log_error(message):
    """Log error messages to the log workbook."""
    logger.error(message)
    error_log_ws.append([message])
    wb_log.save(LOG_XLSX_PATH)

def log_success(message):
    """Log success messages to the log workbook."""
    logger.info(message)
    success_log_ws.append([message])
    wb_log.save(LOG_XLSX_PATH)

class Worker(QThread):
    """Worker thread to handle tasks in the background."""
    progress_signal = pyqtSignal(int)
    error_signal = pyqtSignal(str)
    finished = pyqtSignal()

    def __init__(self, task, *args, **kwargs):
        super().__init__()
        self.task = task
        self.args = args
        self.kwargs = kwargs

    def run(self):
        try:
            self.task(*self.args, **self.kwargs, progress_callback=self.update_progress)
        except Exception as e:
            self.error_signal.emit(str(e))
        self.finished.emit()

    def update_progress(self, progress_value):
        self.progress_signal.emit(progress_value)

async def copy_file_async(src_file, dest_file, semaphore, progress_callback=None):
    """Asynchronously copy files with a semaphore to limit concurrency."""
    async with semaphore:
        try:
            if os.path.isfile(src_file):
                async with aiofiles.open(src_file, 'rb') as src, aiofiles.open(dest_file, 'wb') as dest:
                    await dest.write(await src.read())
                os.chmod(dest_file, stat.S_IWRITE)
                log_success(f"Copied {src_file} to {dest_file}")
            else:
                log_error(f"File not found: {src_file}")
        except Exception as e:
            log_error(f"Error copying {src_file} to {dest_file}: {e}")

        if progress_callback:
            progress_callback(1)

async def copy_entity_files(workbook, version_path, entity_type, progress_callback=None):
    """Copy and sort entity files based on the workbook and entity type."""
    logger.debug(f"Starting copy_entity_files with version_path: {version_path}, entity_type: {entity_type}")
    sheet_name = entity_type
    version_name = os.path.basename(version_path)

    if sheet_name not in workbook.sheetnames:
        log_error(f"Sheet {sheet_name} not found in the workbook.")
        return

    ws = workbook[sheet_name]
    semaphore = asyncio.Semaphore(50)
    tasks = []
    total_files = sum(1 for row in ws.iter_rows(min_row=2, values_only=True) for cell in row[2:] if cell)  # Count all files to be copied
    copied_files = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        entity_id = row[0]
        folder_name = row[1]
        if not folder_name:
            log_error(f"Missing Folder_Name in row: {row}")
            continue

        dest_dir = os.path.join(r"B:\\Kathana-Out\\Sorted", version_name, entity_type, folder_name)
        ensure_directory_exists(dest_dir)
        files_copied = False

        # Copy Mesh files
        for mesh_file in row[2:6]:  # Mesh1 to Mesh4
            if mesh_file:
                src_file = os.path.join(version_path, "resource", "object", entity_type, "Mesh", mesh_file)
                dest_file = os.path.join(dest_dir, mesh_file)
                tasks.append(copy_file_async(src_file, dest_file, semaphore, progress_callback))
                files_copied = True

        # Copy Animation files
        for ani_file in row[6:]:  # Ani1 to Ani70
            if ani_file:
                src_file = os.path.join(version_path, "resource", "object", entity_type, "Ani", ani_file)
                dest_file = os.path.join(dest_dir, ani_file)
                tasks.append(copy_file_async(src_file, dest_file, semaphore, progress_callback))
                files_copied = True

        if not files_copied:
            shutil.rmtree(dest_dir)
            log_error(f"Removed empty directory: {dest_dir}")

    await asyncio.gather(*tasks)

def copy_and_sort_files(version_path, entity_type, progress_callback=None):
    """Copy and sort files for a specific entity type."""
    logger.debug(f"Initiating copy_and_sort_files for {entity_type} from {version_path}")
    logger.info(f"Copying and sorting {entity_type} files from {version_path}...")
    wb = openpyxl.load_workbook(ENTITY_XLSX_PATH)

    start_time = time.time()

    asyncio.run(copy_entity_files(wb, version_path, entity_type, progress_callback=progress_callback))

    end_time = time.time()
    elapsed_time = end_time - start_time
    logger.info(f"{entity_type} files copied and sorted. Time elapsed: {str(timedelta(seconds=elapsed_time))}")

def copy_and_sort_all_files(version_path, progress_callback=None):
    """Copy and sort files for all entity types."""
    copy_and_sort_files(version_path, 'PC', progress_callback=progress_callback)
    copy_and_sort_files(version_path, 'NPC', progress_callback=progress_callback)
    copy_and_sort_files(version_path, 'Monster', progress_callback=progress_callback)

def generate_combined_fbx_batch_file(version_path, progress_callback=None):
    """Generate a combined FBX batch file for all entity types."""
    logger.debug(f"Generating combined FBX batch file for {version_path}")
    batch_commands = []
    generate_fbx_files(version_path, 'PC', batch_commands=batch_commands, progress_callback=progress_callback)
    generate_fbx_files(version_path, 'NPC', batch_commands=batch_commands, progress_callback=progress_callback)
    generate_fbx_files(version_path, 'Monster', batch_commands=batch_commands, progress_callback=progress_callback)

    combined_batch_file_path = os.path.join(r"B:\\Kathana-Out\\Sorted", os.path.basename(version_path), "generate_all_fbx.bat")

    with open(combined_batch_file_path, 'w') as batch_file:
        for command in batch_commands:
            batch_file.write(command + '\n')

    logger.info(f"Combined batch script for generating all entity FBX files created at {combined_batch_file_path}")

def generate_fbx_files(version_path, entity_type, batch_commands=[], progress_callback=None, generate_batch_only=False):
    """Generate FBX files for a specific entity type."""
    logger.debug(f"Generating FBX files for {entity_type} from {version_path}")
    logger.info(f"Generating {entity_type} FBX files from {version_path}...")

    root_dir = os.path.join(r"B:\\Kathana-Out\\Sorted", os.path.basename(version_path), entity_type)
    fbx_base_dir = os.path.join(r"B:\\Kathana-Out\\FBX", os.path.basename(version_path), entity_type)
    ensure_directory_exists(fbx_base_dir)

    if generate_batch_only:
        batch_file_path = os.path.join(root_dir, f"generate_{entity_type.lower()}_fbx.bat")

        # Ensure the directory for the batch file exists
        ensure_directory_exists(os.path.dirname(batch_file_path))

        with open(batch_file_path, 'w') as batch_file:
            for root, dirs, files in os.walk(root_dir):
                for file in files:
                    if file.endswith(".tmb"):
                        tmb_path = os.path.join(root, file)
                        tab_files = [f for f in files if f.endswith(".tab")]
                        for tab_file in tab_files:
                            tab_path = os.path.join(root, tab_file)
                            output_file = os.path.join(fbx_base_dir, os.path.relpath(tab_path, root_dir)).replace(".tab", ".fbx")
                            ensure_directory_exists(os.path.dirname(output_file))
                            command = f'"{NOESIS_EXE_PATH}" ?cmode "{tmb_path}" "{output_file}" -loadanimsingle "{tab_path}" -export -bakeanimscale -showstats -animbonenamematch -fbxnoextraframe'
                            batch_file.write(command + '\n')

        logger.info(f"Batch script for generating {entity_type} FBX files created at {batch_file_path}")
    else:
        if progress_callback:
            progress_callback(1)

def clean_up():
    """Clean up the generated files and directories."""
    logger.debug("Starting clean up process")
    sorted_path = r"B:\\Kathana-Out\\Sorted"
    fbx_path = r"B:\\Kathana-Out\\FBX"
    if os.path.exists(sorted_path):
        shutil.rmtree(sorted_path)
        logger.info("Cleaned up the kathana-res-sorted folder")
    else:
        logger.info("kathana-res-sorted folder does not exist")
    if os.path.exists(fbx_path):
        shutil.rmtree(fbx_path)
        logger.info("Cleaned up the kathana-res-fbx folder")
    else:
        logger.info("kathana-res-fbx folder does not exist")

class KathanaVersionTool(QWidget):
    """Main application window for the Kathana Version Tool."""

    def __init__(self):
        super().__init__()
        self.worker = None
        self.initUI()

    def initUI(self):
        """Initialize the user interface."""
        self.setWindowTitle('Kathana Version Selector')
        self.setFixedSize(800, 650)
        self.setWindowIcon(QIcon(icon_path))

        QApplication.setFont(QFont("Dotum", 8))

        layout = QVBoxLayout()

        self.banner_label = QLabel(self)
        pixmap = QPixmap(banner_path)
        self.banner_label.setPixmap(pixmap)
        self.banner_label.setScaledContents(True)
        self.banner_label.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        layout.addWidget(self.banner_label)

        self.tab_widget = QTabWidget()
        for version_name, version_path in zip(KATHANA_DISPLAY_NAMES, KATHANA_VERSIONS):
            tab = self.create_version_tab(version_name, version_path)
            self.tab_widget.addTab(tab, version_name)
        layout.addWidget(self.tab_widget)

        self.progress_bar = QProgressBar()
        self.progress_bar.setStyleSheet("""
            QProgressBar {
                border: 2px solid grey;
                border-radius: 5px;
                text-align: center;
            }

            QProgressBar::chunk {
                background-color: orange;
                width: 20px;
            }
        """)
        self.progress_bar.setMaximum(100)
        self.progress_bar.setValue(0)
        layout.addWidget(self.progress_bar)

        self.setLayout(layout)

    def create_version_tab(self, version_name, version_path):
        """Create a tab for a specific Kathana version."""
        tab = QWidget()
        layout = QVBoxLayout()

        button_layout = QHBoxLayout()

        col1_layout = QVBoxLayout()
        self.add_button_row(
            col1_layout, version_path,
            [('Copy and Sort PC Files', 'PC'),
             ('Copy and Sort NPC Files', 'NPC'),
             ('Copy and Sort Monster Files', 'Monster'),
             ('Copy and Sort All Entity Files', 'All')]
        )
        button_layout.addLayout(col1_layout)

        col2_layout = QVBoxLayout()
        self.add_button_row(
            col2_layout, version_path,
            [('Generate PC FBX Files', 'PC'),
             ('Generate NPC FBX Files', 'NPC'),
             ('Generate Monster FBX Files', 'Monster'),
             ('Generate All Entity FBX Files', 'All')]
        )
        button_layout.addLayout(col2_layout)

        col3_layout = QVBoxLayout()
        self.add_button_row(
            col3_layout, version_path,
            [('Generate PC FBX Batch File Only', 'PC', True),
             ('Generate NPC FBX Batch File Only', 'NPC', True),
             ('Generate Monster FBX Batch File Only', 'Monster', True),
             ('Generate All FBX Batch File Only', None)]
        )
        button_layout.addLayout(col3_layout)

        layout.addLayout(button_layout)

        control_buttons_layout = QHBoxLayout()
        self.add_button_row(
            control_buttons_layout, None,
            [('Clean Up', None), ('Stop', None), ('Refresh', None)]
        )
        layout.addLayout(control_buttons_layout)

        tab.setLayout(layout)
        return tab

    def add_button_row(self, parent_layout, version_path, buttons):
        """Add a row of buttons to the specified layout."""
        for label, entity_type, *batch_only in buttons:
            if label == 'Clean Up':
                button = QPushButton(label)
                button.clicked.connect(self.confirm_clean_up)
            elif label == 'Stop':
                button = QPushButton(label)
                button.clicked.connect(self.stop_processes)
            elif label == 'Refresh':
                button = QPushButton(label)
                button.clicked.connect(self.restart_application)
            else:
                button = QPushButton(label)
                button.clicked.connect(lambda _, v=version_path, e=entity_type, b=batch_only: self.run_task(v, e, *b))
            button.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Preferred)
            parent_layout.addWidget(button)

    def confirm_clean_up(self):
        """Confirm clean up action with the user."""
        reply = QMessageBox.question(self, 'Confirm Clean Up', 'Are you sure you want to clean up? This action cannot be undone.',
                                     QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No, QMessageBox.StandardButton.No)

        if reply == QMessageBox.StandardButton.Yes:
            self.clean_up()

    def run_task(self, version_path, entity_type, generate_batch_only=False):
        """Run a task to copy and sort files or generate FBX files."""
        self.progress_bar.setValue(0)

        progress_callback = self.update_progress

        if entity_type == 'All':
            self.worker = Worker(copy_and_sort_all_files, version_path, progress_callback=progress_callback)
        elif entity_type is None:
            self.worker = Worker(generate_combined_fbx_batch_file, version_path, progress_callback=progress_callback)
        elif generate_batch_only:
            self.worker = Worker(generate_fbx_files, version_path, entity_type, batch_commands=[], progress_callback=progress_callback, generate_batch_only=True)
        else:
            self.worker = Worker(copy_and_sort_files, version_path, entity_type, progress_callback=progress_callback)

        self.worker.progress_signal.connect(self.update_progress)
        self.worker.error_signal.connect(self.display_error)
        self.worker.finished.connect(self.on_task_finished)
        self.worker.start()

    def update_progress(self, progress_value):
        """Update the progress bar."""
        self.progress_bar.setValue(progress_value)

    def display_error(self, error_message):
        """Display error messages."""
        QMessageBox.critical(self, 'Error', error_message)

    def on_task_finished(self):
        """Handle task finished event."""
        self.progress_bar.setValue(100)
        QMessageBox.information(self, 'Task Finished', 'The task has been completed successfully.')
        self.progress_bar.setValue(0)

    def stop_processes(self):
        """Stop all running processes."""
        if self.worker and self.worker.isRunning():
            self.worker.terminate()
            self.worker.wait()
            self.progress_bar.setValue(0)
            QMessageBox.information(self, 'Process Stopped', 'All running processes have been stopped.')

    def restart_application(self):
        """Restart the application."""
        QApplication.quit()
        os.execl(sys.executable, sys.executable, *sys.argv)

if __name__ == '__main__':
    app = QApplication(sys.argv)
    tool = KathanaVersionTool()
    tool.show()
    sys.exit(app.exec())
