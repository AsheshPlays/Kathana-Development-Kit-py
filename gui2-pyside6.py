import sys
import os
import shutil
import openpyxl
import stat
import logging
import asyncio
import aiofiles
from PySide6.QtWidgets import QApplication, QWidget, QVBoxLayout, QHBoxLayout, QPushButton, QLabel, QTabWidget, QSizePolicy, QTextEdit, QSpacerItem, QMessageBox, QMenu, QProgressBar
from PySide6.QtGui import QPixmap, QIcon, QPalette, QColor, QFont, QPainter, QPolygon, QAction
from PySide6.QtCore import Qt, QProcess, QThread, Signal, QPoint, QTimer
from openpyxl import Workbook
from concurrent.futures import ThreadPoolExecutor
from datetime import timedelta
import time
import pygame

# Initialize pygame for sound
pygame.mixer.init()

# Define the base path to resource files
def resource_path(relative_path):
    """ Get the absolute path to the resource, works for dev and for PyInstaller """
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

# Load resources
startup_sound = resource_path('startup.wav')
hover_sound = resource_path('hover.mp3')
click_sound = resource_path('click.mp3')
done_sound = resource_path('done.mp3')
processing_sound = resource_path('processing.mp3')
icon_path = resource_path('asheshicon.png')
banner_path = resource_path('asheshdevkitbanner.png')

# Project Constants
KATHANA_DISPLAY_NAMES = ["Kathana Global", "Kathana 2", "Kathana 3", "Kathana 3.2", "Kathana 4", "Kathana 5.2",
                         "Kathana 6"]
KATHANA_VERSIONS = [r"B:\\Kathana\\Kathana-Global", r"B:\\Kathana\\Kathana2", r"B:\\Kathana\\Kathana3",
                    r"B:\\Kathana\\Kathana3.2", r"B:\\Kathana\\Kathana4", r"B:\\Kathana\\Kathana5.2",
                    r"B:\\Kathana\\Kathana6"]

LOG_XLSX_FILENAME = "KATHANA_LOGS.xlsx"
LOG_XLSX_PATH = os.path.join(os.getcwd(), LOG_XLSX_FILENAME)
ENTITY_XLSX_PATH = r"B:\\Kathana\\Kathana_Entity_PS.xlsx"
NOESIS_EXE_PATH = r"B:\\Kathana\\_Noesis\\Noesis.exe"

# Initialize logging
logging.basicConfig(level=logging.DEBUG, format='%(message)s')
logger = logging.getLogger()

# Initialize the log workbook with sheets for error and success logs
def initialize_log_workbook():
    wb_log = Workbook()
    error_log_ws = wb_log.create_sheet('ERROR_LOGS')
    success_log_ws = wb_log.create_sheet('SUCCESS_LOGS')
    default_sheet = wb_log.active
    wb_log.remove(default_sheet)
    wb_log.save(LOG_XLSX_PATH)
    return wb_log

# Ensure that the specified directory exists, creating it if necessary
def ensure_directory_exists(path):
    if not os.path.exists(path):
        os.makedirs(path)
        logger.debug(f"Created directory: {path}")

ensure_directory_exists(os.path.dirname(LOG_XLSX_PATH))
wb_log = initialize_log_workbook()
error_log_ws = wb_log['ERROR_LOGS']
success_log_ws = wb_log['SUCCESS_LOGS']

# Log error messages to both the console and the log workbook
def log_error(message):
    logger.error(f"[ERROR] {message}")
    error_log_ws.append([message])
    wb_log.save(LOG_XLSX_PATH)

# Log success messages to both the console and the log workbook
def log_success(message):
    logger.info(f"[SUCCESS] {message}")
    success_log_ws.append([message])
    wb_log.save(LOG_XLSX_PATH)

# Asynchronously copy files with a semaphore to limit concurrency
async def copy_file_async(src_file, dest_file, semaphore):
    logger.debug(f"Attempting to copy from {src_file} to {dest_file}")
    async with semaphore:
        if os.path.isfile(src_file):
            try:
                async with aiofiles.open(src_file, 'rb') as src, aiofiles.open(dest_file, 'wb') as dest:
                    await dest.write(await src.read())
                os.chmod(dest_file, stat.S_IWRITE)
                log_success(f"Copied {src_file} to {dest_file}")
            except Exception as e:
                log_error(f"Error copying {src_file} to {dest_file}: {e}")
        else:
            log_error(f"File not found: {src_file}")

# Copy and sort entity files based on the workbook and entity type
async def copy_entity_files(worker, workbook, version_path, entity_type):
    logger.debug(f"Starting copy_entity_files with version_path: {version_path}, entity_type: {entity_type}")
    sheet_name = entity_type
    version_name = os.path.basename(version_path)

    if sheet_name not in workbook.sheetnames:
        log_error(f"Sheet {sheet_name} not found in the workbook.")
        return

    ws = workbook[sheet_name]
    semaphore = asyncio.Semaphore(20)
    tasks = []
    total_rows = ws.max_row - 1
    completed_rows = 0

    async def copy_files():
        nonlocal completed_rows
        for row in ws.iter_rows(min_row=2, values_only=True):
            if worker.stopped:
                break
            entity_id = row[0]
            folder_name = row[1]
            if not folder_name:
                log_error(f"Missing Folder_Name in row: {row}")
                continue

            dest_dir = os.path.join(r"B:\\Kathana-Out\\Sorted", version_name, entity_type, folder_name)
            ensure_directory_exists(dest_dir)
            files_copied = False

            # Copy Mesh files
            for mesh_file in row[2:6]:
                if mesh_file:
                    src_file = os.path.join(version_path, "resource", "object", entity_type, "Mesh", mesh_file)
                    dest_file = os.path.join(dest_dir, mesh_file)
                    tasks.append(copy_file_async(src_file, dest_file, semaphore))
                    files_copied = True

            # Copy Animation files
            for ani_file in row[6:]:
                if ani_file:
                    src_file = os.path.join(version_path, "resource", "object", entity_type, "Ani", ani_file)
                    dest_file = os.path.join(dest_dir, ani_file)
                    tasks.append(copy_file_async(src_file, dest_file, semaphore))
                    files_copied = True

            if not files_copied:
                shutil.rmtree(dest_dir)
                log_error(f"Removed empty directory: {dest_dir}")

            completed_rows += 1
            progress = int((completed_rows / total_rows) * 100)
            worker.progress.emit(progress)
            worker.progress_info.emit(completed_rows, total_rows)

    await copy_files()
    await asyncio.gather(*tasks)

# Copy and sort files for a specific entity type
def copy_and_sort_files(worker, version_path, entity_type):
    logger.debug(f"Initiating copy_and_sort_files for {entity_type} from {version_path}")
    logger.info(f"Copying and sorting {entity_type} files from {version_path}...")
    wb = openpyxl.load_workbook(ENTITY_XLSX_PATH)
    start_time = time.time()

    with ThreadPoolExecutor(max_workers=20) as executor:
        future = executor.submit(asyncio.run, copy_entity_files(worker, wb, version_path, entity_type))
        future.result()

    end_time = time.time()
    elapsed_time = end_time - start_time
    logger.info(f"{entity_type} files copied and sorted. Time elapsed: {str(timedelta(seconds=elapsed_time))}")

# Copy and sort files for all entity types
def copy_and_sort_all_files(worker, version_path):
    copy_and_sort_files(worker, version_path, 'PC')
    copy_and_sort_files(worker, version_path, 'NPC')
    copy_and_sort_files(worker, version_path, 'Monster')

# Generate FBX files for a specific entity type
def generate_fbx_files(worker, version_path, entity_type, generate_batch_only=False, combined_batch=False, batch_commands=[]):
    logger.debug(f"Generating FBX files for {entity_type} from {version_path}")
    logger.info(f"Generating {entity_type} FBX files from {version_path}...")

    root_dir = os.path.join(r"B:\\Kathana-Out\\Sorted", os.path.basename(version_path), entity_type)
    fbx_base_dir = os.path.join(r"B:\\Kathana-Out\\FBX", os.path.basename(version_path), entity_type)
    ensure_directory_exists(fbx_base_dir)

    if combined_batch:
        for root, dirs, files in os.walk(root_dir):
            for file in files:
                if file.endswith(".tmb"):
                    tmb_path = os.path.join(root, file)
                    tab_files = [f for f in files if f.endswith(".tab")]
                    for tab_file in tab_files:
                        tab_path = os.path.join(root, tab_file)
                        output_file = os.path.join(fbx_base_dir, os.path.relpath(tab_path, root_dir)).replace(".tab", ".fbx")
                        ensure_directory_exists(os.path.dirname(output_file))
                        command = f'"{NOESIS_EXE_PATH}" ?cmode "{tmb_path}" "{output_file}" -loadanimsingle "{tab_path}" -fbxnoextraframe'
                        batch_commands.append(command)
    else:
        batch_file_path = os.path.join(root_dir, f"generate_{entity_type.lower()}_fbx.bat")
        ensure_directory_exists(os.path.dirname(batch_file_path))

        with open(batch_file_path, 'w') as batch_file:
            for root, dirs, files in os.walk(root_dir):
                for file in files:
                    if file.endswith(".tmb"):
                        tmb_path = os.path.join(root, file)
                        tab_files = [f for f in files if f.endswith(".tab")]
                        for tab_file in tab_files:
                            tab_path = os.path.join(root, tab_file)
                            output_file = os.path.join(fbx_base_dir, os.path.relpath(tab_path, root_dir)).replace(".tab", ".fbx")
                            ensure_directory_exists(os.path.dirname(output_file))
                            command = f'"{NOESIS_EXE_PATH}" ?cmode "{tmb_path}" "{output_file}" -loadanimsingle "{tab_path}" -fbxnoextraframe'
                            batch_file.write(command + '\n')

        logger.info(f"Batch script for generating {entity_type} FBX files created at {batch_file_path}")
        if not generate_batch_only:
            os.system(f'cmd /c "{batch_file_path}"')
            logger.info(f"{entity_type} FBX files generation complete.")

# Generate a combined FBX batch file for all entity types
def generate_combined_fbx_batch_file(worker, version_path):
    logger.debug(f"Generating combined FBX batch file for {version_path}")
    batch_commands = []
    generate_fbx_files(worker, version_path, 'PC', generate_batch_only=True, combined_batch=True, batch_commands=batch_commands)
    generate_fbx_files(worker, version_path, 'NPC', generate_batch_only=True, combined_batch=True, batch_commands=batch_commands)
    generate_fbx_files(worker, version_path, 'Monster', generate_batch_only=True, combined_batch=True, batch_commands=batch_commands)

    combined_batch_file_path = os.path.join(r"B:\\Kathana-Out\\Sorted", os.path.basename(version_path), "generate_all_fbx.bat")
    with open(combined_batch_file_path, 'w') as batch_file:
        for command in batch_commands:
            batch_file.write(command + '\n')

    logger.info(f"Combined batch script for generating all entity FBX files created at {combined_batch_file_path}")

# Clean specific files for a given version and entity type
def clean_specific_files(worker, version, entity_type):
    sorted_path = os.path.join(r"B:\\Kathana-Out\\Sorted", os.path.basename(version), entity_type)
    fbx_path = os.path.join(r"B:\\Kathana-Out\\FBX", os.path.basename(version), entity_type)
    if os.path.exists(sorted_path):
        shutil.rmtree(sorted_path)
        log_success(f"Cleaned {entity_type} files for version {version} from kathana-res-sorted folder")
    else:
        log_error(f"{sorted_path} does not exist")
    if os.path.exists(fbx_path):
        shutil.rmtree(fbx_path)
        log_success(f"Cleaned {entity_type} files for version {version} from kathana-res-fbx folder")
    else:
        log_error(f"{fbx_path} does not exist")

# Clean up all generated files and directories
def clean_up(worker):
    logger.debug("Starting clean up process")
    sorted_path = r"B:\\Kathana-Out\\Sorted"
    fbx_path = r"B:\\Kathana-Out\\FBX"
    if os.path.exists(sorted_path):
        shutil.rmtree(sorted_path)
        logger.info("Cleaned up the kathana-res-sorted folder")
    else:
        logger.info("kathana-res-sorted folder does not exist")
    if os.path.exists(fbx_path):
        shutil.rmtree(fbx_path)
        logger.info("Cleaned up the kathana-res-fbx folder")
    else:
        logger.info("kathana-res-fbx folder does not exist")

class Worker(QThread):
    """Worker thread to handle tasks in the background."""
    progress = Signal(int)
    progress_info = Signal(int, int)
    output = Signal(str)
    error = Signal(str)
    finished = Signal()

    def __init__(self, task, *args, **kwargs):
        super().__init__()
        self.task = task
        self.args = args
        self.kwargs = kwargs
        self.stopped = False

    def run(self):
        try:
            self.task(self, *self.args, **self.kwargs)
        except Exception as e:
            self.error.emit(str(e))
        self.finished.emit()

    def stop(self):
        self.stopped = True

class ArrowButton(QPushButton):
    """Custom button to display an arrow pointing left or right with sound effects."""
    def __init__(self, direction, hover_sound, click_sound, parent=None):
        super().__init__(parent)
        self.direction = direction
        self.hover_sound = hover_sound
        self.click_sound = click_sound
        self.setFixedSize(30, 30)
        self.setStyleSheet("background-color: transparent; border: none;")

        # Load sounds
        try:
            self.hover_sound_effect = pygame.mixer.Sound(self.hover_sound)
            self.click_sound_effect = pygame.mixer.Sound(self.click_sound)
        except pygame.error as e:
            logger.error(f"Error loading sound: {e}")
            self.hover_sound_effect = None
            self.click_sound_effect = None

    def paintEvent(self, event):
        painter = QPainter(self)
        painter.setRenderHint(QPainter.RenderHint.Antialiasing)
        painter.setBrush(QColor("red"))
        painter.setPen(Qt.PenStyle.NoPen)

        if self.direction == "left":
            points = [QPoint(self.width(), 0), QPoint(0, self.height() // 2), QPoint(self.width(), self.height())]
        elif self.direction == "right":
            points = [QPoint(0, 0), QPoint(self.width(), self.height() // 2), QPoint(0, self.height())]

        triangle = QPolygon(points)
        painter.drawPolygon(triangle)

    def enterEvent(self, event):
        if self.hover_sound_effect:
            self.hover_sound_effect.play()
        super().enterEvent(event)

    def mousePressEvent(self, event):
        if self.click_sound_effect:
            self.click_sound_effect.play()
        super().mousePressEvent(event)

class SoundButton(QPushButton):
    """Custom button with sound effects and rounded corners."""
    def __init__(self, label, hover_sound, click_sound, parent=None):
        super().__init__(label, parent)
        self.hover_sound = hover_sound
        self.click_sound = click_sound
        self.setStyleSheet("""
            QPushButton {
                background-color: #6272a4;
                color: #f8f8f2;
                border: 2px solid #bd93f9;
                height: 30px;
                border-radius: 10px;
                transition: transform 0.2s;
            }
            QPushButton:hover {
                background-color: #bd93f9;
                transform: scale(1.05);
            }
            QPushButton:pressed {
                background-color: #44475a;
            }
        """)

        # Load sounds
        try:
            self.hover_sound_effect = pygame.mixer.Sound(self.hover_sound)
            self.click_sound_effect = pygame.mixer.Sound(self.click_sound)
        except pygame.error as e:
            logger.error(f"Error loading sound: {e}")
            self.hover_sound_effect = None
            self.click_sound_effect = None

    def enterEvent(self, event):
        if self.hover_sound_effect:
            self.hover_sound_effect.play()
        super().enterEvent(event)

    def mousePressEvent(self, event):
        if self.click_sound_effect:
            self.click_sound_effect.play()
        super().mousePressEvent(event)

class SignalHandler(logging.Handler):
    """Custom logging handler to emit log records to a Qt signal."""
    def __init__(self, signal):
        super().__init__()
        self.signal = signal

    def emit(self, record):
        log_entry = self.format(record)
        self.signal.emit(log_entry)

class KathanaVersionTool(QWidget):
    """Main application window for the Kathana Version Tool."""
    append_log = Signal(str)

    def __init__(self):
        super().__init__()
        self.selected_index = 0
        self.selected_version = KATHANA_VERSIONS[self.selected_index]
        self.process = QProcess(self)
        self.version_set = False
        self.worker = None
        self.initUI()

        self.logger = logging.getLogger()
        self.logger.setLevel(logging.DEBUG)
        handler = SignalHandler(self.append_log)
        handler.setFormatter(logging.Formatter('%(message)s'))
        self.logger.addHandler(handler)
        self.append_log.connect(self.append_output)

        # Play startup sound
        pygame.mixer.Sound(startup_sound).play()

    def initUI(self):
        """Initialize the user interface."""
        self.setWindowTitle('Kathana Version Selector')
        self.setFixedSize(800, 700)
        self.setWindowIcon(QIcon(icon_path))

        QApplication.setFont(QFont("Dotum", 8))

        layout = QVBoxLayout()

        self.banner_label = QLabel(self)
        pixmap = QPixmap(banner_path)
        self.banner_label.setPixmap(pixmap)
        self.banner_label.setScaledContents(True)
        self.banner_label.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        layout.addWidget(self.banner_label)

        self.tab_widget = QTabWidget()
        for i, version in enumerate(KATHANA_DISPLAY_NAMES):
            self.tab_widget.addTab(QWidget(), version)
        self.tab_widget.setCurrentIndex(0)
        self.tab_widget.currentChanged.connect(self.tab_changed)
        layout.addWidget(self.tab_widget)

        button_layout = QHBoxLayout()

        set_version_btn = SoundButton('Set Version', hover_sound, click_sound, self)
        set_version_btn.clicked.connect(self.set_version)
        button_layout.addWidget(set_version_btn)

        layout.addLayout(button_layout)

        buttons_layout = QHBoxLayout()

        col1_layout = QVBoxLayout()
        self.add_button_row(
            col1_layout, [('Copy and Sort PC Files', lambda: self.run_task('PC')),
                          ('Copy and Sort NPC Files', lambda: self.run_task('NPC')),
                          ('Copy and Sort Monster Files', lambda: self.run_task('Monster')),
                          ('Copy and Sort All Entity Files', lambda: self.run_task('All'))]
        )
        buttons_layout.addLayout(col1_layout)

        col2_layout = QVBoxLayout()
        self.add_button_row(
            col2_layout, [('Generate PC FBX Files', lambda: self.run_fbx_task('PC')),
                          ('Generate NPC FBX Files', lambda: self.run_fbx_task('NPC')),
                          ('Generate Monster FBX Files', lambda: self.run_fbx_task('Monster')),
                          ('Generate All Entity FBX Files', lambda: self.run_fbx_task('All'))]
        )
        buttons_layout.addLayout(col2_layout)

        col3_layout = QVBoxLayout()
        self.add_button_row(
            col3_layout, [('Generate PC FBX Batch File Only', lambda: self.run_fbx_task('PC', True)),
                          ('Generate NPC FBX Batch File Only', lambda: self.run_fbx_task('NPC', True)),
                          ('Generate Monster FBX Batch File Only', lambda: self.run_fbx_task('Monster', True)),
                          ('Generate All FBX Batch File Only', self.run_combined_fbx_batch_file)]
        )
        buttons_layout.addLayout(col3_layout)

        layout.addLayout(buttons_layout)

        control_buttons_layout = QHBoxLayout()
        self.add_button_row(
            control_buttons_layout,
            [('Clean Up', self.clean_up), ('Stop', self.stop_processes), ('Refresh', self.restart_application)]
        )
        layout.addLayout(control_buttons_layout)

        console_section_layout = QVBoxLayout()

        console_label = QLabel('Console')
        console_label.setFont(QFont("Dotum", 10, QFont.Weight.Bold))
        console_section_layout.addWidget(console_label)

        console_layout = QHBoxLayout()

        self.console_output = QTextEdit()
        self.console_output.setReadOnly(True)
        console_palette = self.console_output.palette()
        console_palette.setColor(QPalette.ColorRole.Base, QColor(30, 32, 40, 255))
        self.console_output.setPalette(console_palette)
        self.console_output.setFixedHeight(130)
        console_layout.addWidget(self.console_output)

        console_buttons_layout = QVBoxLayout()
        clear_console_btn = SoundButton('Clear \nConsole', hover_sound, click_sound, self)
        clear_console_btn.setStyleSheet("""
            QPushButton {
                background-color: #ff5555;
                color: white;
                border: 2px solid #ff79c6;
                border-radius: 10px;
            }
            QPushButton:hover {
                background-color: #ff79c6;
            }
        """)
        clear_console_btn.setFont(QFont("Dotum", 10))
        clear_console_btn.clicked.connect(self.clear_console)
        clear_console_btn.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        clear_console_btn.setFixedWidth(125)
        clear_console_btn.setFixedHeight(125)
        console_buttons_layout.addWidget(clear_console_btn)

        console_layout.addLayout(console_buttons_layout)
        console_section_layout.addLayout(console_layout)
        layout.addLayout(console_section_layout)

        self.progress_bar = QProgressBar(self)
        self.progress_bar.setFixedHeight(10)
        self.progress_bar.setStyleSheet("""
            QProgressBar {
                background-color: #44475a;
                border: 2px solid #bd93f9;
                border-radius: 5px;
                text-align: center;
            }
            QProgressBar::chunk {
                background-color: #bd93f9;
            }
        """)
        self.progress_label_left = QLabel('0 / 0')
        self.progress_label_right = QLabel('0%')
        progress_layout = QHBoxLayout()
        progress_layout.addWidget(self.progress_label_left)
        progress_layout.addWidget(self.progress_bar)
        progress_layout.addWidget(self.progress_label_right)
        layout.addLayout(progress_layout)

        author_version_layout = QHBoxLayout()
        author_label = QLabel('Ashesh Development © 2024')
        author_label.setFont(QFont("Dotum", 8))
        version_label = QLabel('Version: 1.1.3')
        version_label.setFont(QFont("Dotum", 8))
        author_version_layout.addWidget(author_label)
        author_version_layout.addSpacerItem(QSpacerItem(40, 20, QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Minimum))
        author_version_layout.addWidget(version_label)
        layout.addLayout(author_version_layout)

        self.setLayout(layout)

        self.process.setProcessChannelMode(QProcess.ProcessChannelMode.MergedChannels)
        self.process.readyReadStandardOutput.connect(self.on_readyReadStandardOutput)
        self.process.readyReadStandardError.connect(self.on_readyReadStandardError)
        self.process.finished.connect(self.on_task_finished)

        self.timer = QTimer()
        self.timer.timeout.connect(self.update_progress_labels)

    def add_button_row(self, parent_layout, buttons):
        """Add a row of buttons to the specified layout."""
        for label, callback in buttons:
            if label == 'Clean Up':
                button = SoundButton(label, hover_sound, click_sound)
                button.setContextMenuPolicy(Qt.CustomContextMenu)
                button.customContextMenuRequested.connect(self.show_clean_up_menu)
                button.clicked.connect(self.confirm_clean_up)
                button.setStyleSheet("""
                    QPushButton {
                        background-color: #282a36;
                        color: white;
                        border: 2px solid #ff5555;
                        height: 30px;  
                        border-radius: 10px;  
                        transition: transform 0.2s;
                    }
                    QPushButton:hover {
                        background-color: #ff79c6;
                        transform: scale(1.05);  
                    }
                    QPushButton:pressed {
                        background-color: #44475a;
                    }
                """)
            else:
                button = SoundButton(label, hover_sound, click_sound)
                button.clicked.connect(callback)
            button.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Preferred)
            parent_layout.addWidget(button)

    def show_clean_up_menu(self, pos):
        """Show the context menu for the Clean Up button."""
        button = self.sender()
        menu = QMenu(self)
        for version in KATHANA_VERSIONS:
            version_menu = QMenu(os.path.basename(version), self)
            for entity in ['PC', 'NPC', 'Monster', 'All']:
                action = QAction(f"Clean {entity}", self)
                action.triggered.connect(lambda checked, v=version, e=entity: self.clean_specific_files(v, e))
                version_menu.addAction(action)
            menu.addMenu(version_menu)
        global_pos = button.mapToGlobal(pos)
        menu.exec(global_pos)

    def clean_specific_files(self, version, entity_type):
        """Run a task to clean specific files for a given version and entity type."""
        self.start_processing_sound()
        self.progress_bar.setValue(0)
        self.disable_buttons()
        self.worker = Worker(clean_specific_files, version, entity_type)
        self.worker.output.connect(self.append_output)
        self.worker.error.connect(self.append_error)
        self.worker.finished.connect(self.on_task_finished)
        self.worker.progress.connect(self.update_progress)
        self.worker.progress_info.connect(self.update_progress_info)
        self.worker.start()

    def confirm_clean_up(self):
        """Confirm clean up action with the user."""
        reply = QMessageBox.question(self, 'Confirm Clean Up', 'Are you sure you want to clean up? This action cannot be undone.',
                                     QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No, QMessageBox.StandardButton.No)

        if reply == QMessageBox.StandardButton.Yes:
            self.clean_up()

    def tab_changed(self, index):
        """Update the selected Kathana version based on the selected tab."""
        self.selected_index = index
        self.selected_version = KATHANA_VERSIONS[self.selected_index]
        self.append_output(f'Selected version set to: {KATHANA_DISPLAY_NAMES[self.selected_index]}')

    def set_version(self):
        """Set the selected Kathana version and log the action."""
        self.version_set = True
        self.append_output(f'Selected version set to: {KATHANA_DISPLAY_NAMES[self.selected_index]}')

    def run_task(self, entity_type):
        """Run a task to copy and sort files for the specified entity type."""
        if self.selected_version:
            self.start_processing_sound()
            self.progress_bar.setValue(0)
            self.disable_buttons()
            if entity_type == 'All':
                self.worker = Worker(copy_and_sort_all_files, self.selected_version)
            else:
                self.worker = Worker(copy_and_sort_files, self.selected_version, entity_type)
            self.worker.output.connect(self.append_output)
            self.worker.error.connect(self.append_error)
            self.worker.finished.connect(self.on_task_finished)
            self.worker.progress.connect(self.update_progress)
            self.worker.progress_info.connect(self.update_progress_info)
            self.worker.start()
            self.timer.start(100)
        else:
            self.append_error('Error: Please choose a Kathana version first.')

    def run_fbx_task(self, entity_type, generate_batch_only=False):
        """Run a task to generate FBX files for the specified entity type."""
        if self.selected_version:
            self.start_processing_sound()
            self.progress_bar.setValue(0)
            self.disable_buttons()
            self.worker = Worker(generate_fbx_files, self.selected_version, entity_type, generate_batch_only)
            self.worker.output.connect(self.append_output)
            self.worker.error.connect(self.append_error)
            self.worker.finished.connect(self.on_task_finished)
            self.worker.progress.connect(self.update_progress)
            self.worker.progress_info.connect(self.update_progress_info)
            self.worker.start()
            self.timer.start(100)
        else:
            self.append_error('Error: Please choose a Kathana version first.')

    def run_combined_fbx_batch_file(self):
        """Run a task to generate combined FBX batch files."""
        if self.selected_version:
            self.start_processing_sound()
            self.progress_bar.setValue(0)
            self.disable_buttons()
            self.worker = Worker(generate_combined_fbx_batch_file, self.selected_version)
            self.worker.output.connect(self.append_output)
            self.worker.error.connect(self.append_error)
            self.worker.finished.connect(self.on_task_finished)
            self.worker.progress.connect(self.update_progress)
            self.worker.progress_info.connect(self.update_progress_info)
            self.worker.start()
            self.timer.start(100)
        else:
            self.append_error('Error: Please choose a Kathana version first.')

    def clean_up(self):
        """Run a task to clean up generated files."""
        self.start_processing_sound()
        self.progress_bar.setValue(0)
        self.disable_buttons()
        self.worker = Worker(clean_up)
        self.worker.output.connect(self.append_output)
        self.worker.error.connect(self.append_error)
        self.worker.finished.connect(self.on_task_finished)
        self.worker.progress.connect(self.update_progress)
        self.worker.progress_info.connect(self.update_progress_info)
        self.worker.start()
        self.timer.start(100)
        QMessageBox.information(self, 'Clean Up', 'Clean up process has been completed.')

    def stop_processes(self):
        """Stop all running processes."""
        if self.worker:
            self.worker.stop()
        self.append_output('Process Stopped: All running processes have been stopped.')
        self.enable_buttons()

    def restart_application(self):
        """Restart the application."""
        QApplication.quit()
        os.execl(sys.executable, sys.executable, *sys.argv)

    def clear_console(self):
        """Clear the console output."""
        self.console_output.clear()

    def append_output(self, text):
        """Append output text to the console."""
        self.console_output.setTextColor(QColor("orange"))
        self.console_output.append(text)
        self.console_output.setTextColor(QColor("black"))

    def append_error(self, text):
        """Append error text to the console."""
        self.console_output.setTextColor(QColor("red"))
        self.console_output.append(f"ERROR: {text}")
        self.console_output.setTextColor(QColor("black"))

    def on_task_finished(self):
        """Handle task finished event."""
        self.stop_processing_sound()
        self.enable_buttons()
        self.console_output.setTextColor(QColor("green"))
        self.console_output.append(f"Task Finished: {self.worker.task.__name__} has been completed successfully.")
        self.console_output.setTextColor(QColor("black"))
        pygame.mixer.Sound(done_sound).play()
        self.progress_bar.setValue(100)
        self.timer.stop()

    def on_readyReadStandardOutput(self):
        """Handle ready read standard output event."""
        text = self.process.readAllStandardOutput().data().decode()
        self.append_output(text)

    def on_readyReadStandardError(self):
        """Handle ready read standard error event."""
        text = self.process.readAllStandardError().data().decode()
        self.append_error(text)

    def start_processing_sound(self):
        """Start the processing sound."""
        pygame.mixer.Sound(processing_sound).play()

    def stop_processing_sound(self):
        """Stop the processing sound."""
        pygame.mixer.stop()

    def disable_buttons(self):
        """Disable buttons during task execution."""
        for btn in self.findChildren(QPushButton):
            if btn.text() != 'Stop':
                btn.setDisabled(True)

    def enable_buttons(self):
        """Enable buttons after task completion."""
        for btn in self.findChildren(QPushButton):
            btn.setDisabled(False)

    def update_progress(self, value):
        """Update the progress bar value."""
        self.progress_bar.setValue(value)

    def update_progress_info(self, processed, total):
        """Update the progress information labels."""
        self.progress_label_left.setText(f'{processed} / {total}')
        self.progress_label_right.setText(f'{self.progress_bar.value()}%')

    def update_progress_labels(self):
        """Update the progress labels and bar every 100ms."""
        if self.worker:
            self.update_progress_info(self.worker.progress_info_value[0], self.worker.progress_info_value[1])

if __name__ == '__main__':
    app = QApplication(sys.argv)
    # Apply PyDracula theme
    app.setStyle("Fusion")
    palette = QPalette()
    palette.setColor(QPalette.Window, QColor(40, 42, 54))
    palette.setColor(QPalette.WindowText, QColor(255, 255, 255))
    palette.setColor(QPalette.Base, QColor(40, 42, 54))
    palette.setColor(QPalette.AlternateBase, QColor(50, 52, 65))
    palette.setColor(QPalette.ToolTipBase, QColor(255, 255, 255))
    palette.setColor(QPalette.ToolTipText, QColor(255, 255, 255))
    palette.setColor(QPalette.Text, QColor(255, 255, 255))
    palette.setColor(QPalette.Button, QColor(68, 71, 90))
    palette.setColor(QPalette.ButtonText, QColor(255, 255, 255))
    palette.setColor(QPalette.BrightText, QColor(255, 0, 0))
    palette.setColor(QPalette.Link, QColor(85, 170, 255))
    palette.setColor(QPalette.Highlight, QColor(85, 170, 255))
    palette.setColor(QPalette.HighlightedText, QColor(255, 255, 255))
    app.setPalette(palette)

    tool = KathanaVersionTool()
    tool.show()
    sys.exit(app.exec())
