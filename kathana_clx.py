import sys
import os
import shutil
import openpyxl
import stat
from termcolor import colored
import logging
from openpyxl import Workbook
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime, timedelta
import time
import asyncio
import aiofiles

logging.basicConfig(level = logging.DEBUG, format = '%(message)s')
logger = logging.getLogger()

KATHANA_VERSIONS = [r"B:\\Kathana\\Kathana-Global", r"B:\\Kathana\\Kathana2", r"B:\\Kathana\\Kathana3",
                    r"B:\\Kathana\\Kathana3.2", r"B:\\Kathana\\Kathana4", r"B:\\Kathana\\Kathana5.2",
                    r"B:\\Kathana\\Kathana6"]

ENTITY_XLSX_PATH = r"B:\\Kathana\\Kathana_Entity.xlsx"
LOG_XLSX_FILENAME = f"KATHANA_LOGS.xlsx"
LOG_XLSX_PATH = os.path.join(os.getcwd(), LOG_XLSX_FILENAME)
NOESIS_EXE_PATH = r"B:\\Kathana\\_Noesis\\Noesis.exe"

BANNER = """
 █████╗ ███████╗██╗  ██╗███████╗███████╗██╗  ██╗    ██████╗ ███████╗██╗   ██╗███████╗██╗      ██████╗ ██████╗ ███╗   ███╗███████╗███╗   ██╗████████╗
██╔══██╗██╔════╝██║  ██║██╔════╝██╔════╝██║  ██║    ██╔══██╗██╔════╝██║   ██║██╔════╝██║     ██╔═══██╗██╔══██╗████╗ ████║██╔════╝████╗  ██║╚══██╔══╝
███████║███████╗███████║█████╗  ███████╗███████║    ██║  ██║█████╗  ██║   ██║█████╗  ██║     ██║   ██║██████╔╝██╔████╔██║█████╗  ██╔██╗ ██║   ██║
██╔══██║╚════██║██╔══██║██╔══╝  ╚════██║██╔══██║    ██║  ██║██╔══╝  ╚██╗ ██╔╝██╔══╝  ██║     ██║   ██║██╔═══╝ ██║╚██╔╝██║██╔══╝  ██║╚██╗██║   ██║
██║  ██║███████║██║  ██║███████╗███████║██║  ██║    ██████╔╝███████╗ ╚████╔╝ ███████╗███████╗╚██████╔╝██║     ██║ ╚═╝ ██║███████╗██║ ╚████║   ██║
╚═╝  ╚═╝╚══════╝╚═╝  ╚═╝╚══════╝╚══════╝╚═╝  ╚═╝    ╚═════╝ ╚══════╝  ╚═══╝  ╚══════╝╚══════╝ ╚═════╝ ╚═╝     ╚═╝     ╚═╝╚══════╝╚═╝  ╚═══╝   ╚═╝
"""


def initialize_log_workbook():
	wb_log = Workbook()
	error_log_ws = wb_log.create_sheet('ERROR_LOGS')
	success_log_ws = wb_log.create_sheet('SUCCESS_LOGS')
	default_sheet = wb_log.active
	wb_log.remove(default_sheet)
	wb_log.save(LOG_XLSX_PATH)
	return wb_log


wb_log = initialize_log_workbook()
error_log_ws = wb_log['ERROR_LOGS']
success_log_ws = wb_log['SUCCESS_LOGS']


def log_error(message):
	logger.error(colored(message, 'red'))
	error_log_ws.append([message])
	wb_log.save(LOG_XLSX_PATH)


def log_success(message):
	logger.info(colored(message, 'green'))
	success_log_ws.append([message])
	wb_log.save(LOG_XLSX_PATH)


def ensure_directory_exists(path):
	if not os.path.exists(path):
		os.makedirs(path)
		logger.debug(colored(f"Created directory: {path}", 'green'))


async def copy_file_async(src_file, dest_file, semaphore):
	logger.debug(
			colored(f"Entering copy_file_async function with src_file: {src_file} and dest_file: {dest_file}", 'cyan')
			)
	async with semaphore:
		if os.path.isfile(src_file):
			try:
				async with aiofiles.open(src_file, 'rb') as src, aiofiles.open(dest_file, 'wb') as dest:
					await dest.write(await src.read())
				os.chmod(dest_file, stat.S_IWRITE)
				log_success(f"Copied {src_file} to {dest_file}")
			except Exception as e:
				log_error(f"Error copying {src_file} to {dest_file}: {e}")
		else:
			log_error(f"File not found: {src_file}")


async def copy_pc_files(workbook, version_path):
	logger.debug(colored(f"Entering copy_pc_files function with version_path: {version_path}", 'cyan'))
	mesh_sheet = "PC_Mesh"
	ani_sheet = "PC_Ani"
	version_name = os.path.basename(version_path)
	
	semaphore = asyncio.Semaphore(50)
	
	if mesh_sheet in workbook.sheetnames:
		mesh_ws = workbook[mesh_sheet]
		logger.debug(colored(f"Found sheet: {mesh_sheet}", 'cyan'))
		
		pc_tasks = []
		for row in mesh_ws.iter_rows(min_row = 2, values_only = True):
			pc_id, pc_code, pc_mesh_file = row
			if not pc_code or not pc_mesh_file:
				log_error(f"Missing data in row: {row}")
				continue
			dest_dir = os.path.join(r"B:\\Kathana-Out\Sorted", version_name, "PC", pc_code)
			ensure_directory_exists(dest_dir)
			src_file = os.path.join(version_path, "resource", "object", "PC", "Mesh", pc_mesh_file)
			dest_file = os.path.join(dest_dir, pc_mesh_file)
			pc_tasks.append(copy_file_async(src_file, dest_file, semaphore))
		
		await asyncio.gather(*pc_tasks)
	
	if ani_sheet in workbook.sheetnames:
		ani_ws = workbook[ani_sheet]
		logger.debug(colored(f"Found sheet: {ani_sheet}", 'cyan'))
		
		ani_tasks = []
		for col_idx, col in enumerate(ani_ws.iter_cols(min_col = 2, values_only = True), start = 2):
			pc_code = ani_ws.cell(row = 1, column = col_idx).value
			if pc_code:
				dest_dir = os.path.join(r"B:\\Kathana-Out\Sorted", version_name, "PC", pc_code)
				ensure_directory_exists(dest_dir)
				
				files_copied = False
				for ani_file in col:
					if ani_file:
						src_file = os.path.join(version_path, "resource", "object", "PC", "Ani", ani_file)
						dest_file = os.path.join(dest_dir, ani_file)
						ani_tasks.append(copy_file_async(src_file, dest_file, semaphore))
						files_copied = True
				
				if not files_copied:
					shutil.rmtree(dest_dir)
					log_error(f"Removed empty directory: {dest_dir}")
		
		await asyncio.gather(*ani_tasks)


async def copy_npc_files(workbook, version_path):
	logger.debug(colored(f"Entering copy_npc_files function with version_path: {version_path}", 'cyan'))
	mesh_sheet = "NPC_Mesh"
	ani_sheet = "NPC_Ani"
	version_name = os.path.basename(version_path)
	
	semaphore = asyncio.Semaphore(50)
	
	if mesh_sheet in workbook.sheetnames:
		mesh_ws = workbook[mesh_sheet]
		logger.debug(colored(f"Found sheet: {mesh_sheet}", 'cyan'))
		
		npc_tasks = []
		for row in mesh_ws.iter_rows(min_row = 2, values_only = True):
			npc_id, npc_code, npc_mesh_file = row
			if not npc_code or not npc_mesh_file:
				log_error(f"Missing data in row: {row}")
				continue
			dest_dir = os.path.join(r"B:\\Kathana-Out\Sorted", version_name, "NPC", npc_code)
			ensure_directory_exists(dest_dir)
			src_file = os.path.join(version_path, "resource", "object", "NPC", "Mesh", npc_mesh_file)
			dest_file = os.path.join(dest_dir, npc_mesh_file)
			npc_tasks.append(copy_file_async(src_file, dest_file, semaphore))
		
		await asyncio.gather(*npc_tasks)
	
	if ani_sheet in workbook.sheetnames:
		ani_ws = workbook[ani_sheet]
		logger.debug(colored(f"Found sheet: {ani_sheet}", 'cyan'))
		
		ani_tasks = []
		for row in ani_ws.iter_rows(min_row = 2, values_only = True):
			npc_code = row[1]
			if not npc_code:
				log_error(f"Missing NPC_Code in row: {row}")
				continue
			dest_dir = os.path.join(r"B:\\Kathana-Out\\Sorted", version_name, "NPC", npc_code)
			ensure_directory_exists(dest_dir)
			
			files_copied = False
			for header, ani_file in zip(ani_ws[1][2:], row[2:]):
				if ani_file:
					src_file = os.path.join(version_path, "resource", "object", "NPC", "Ani", ani_file)
					dest_file = os.path.join(dest_dir, ani_file)
					ani_tasks.append(copy_file_async(src_file, dest_file, semaphore))
					files_copied = True
			
			if not files_copied:
				shutil.rmtree(dest_dir)
				log_error(f"Removed empty directory: {dest_dir}")
		
		await asyncio.gather(*ani_tasks)


async def copy_monster_files(workbook, version_path):
	logger.debug(colored(f"Entering copy_monster_files function with version_path: {version_path}", 'cyan'))
	mesh_sheet = "Monster_Mesh"
	ani_sheet = "Monster_Ani"
	version_name = os.path.basename(version_path)
	
	semaphore = asyncio.Semaphore(50)
	
	if mesh_sheet in workbook.sheetnames:
		mesh_ws = workbook[mesh_sheet]
		logger.debug(colored(f"Found sheet: {mesh_sheet}", 'cyan'))
		
		monster_tasks = []
		for row in mesh_ws.iter_rows(min_row = 2, values_only = True):
			monster_id, monster_code, *mesh_files = row
			if not monster_code:
				log_error(f"Missing Monster_Code in row: {row}")
				continue
			dest_dir = os.path.join(r"B:\\Kathana-Out\Sorted", version_name, "Monster", monster_code)
			ensure_directory_exists(dest_dir)
			files_copied = False
			for mesh_file in mesh_files:
				if mesh_file:
					src_file = os.path.join(version_path, "resource", "object", "Monster", "Mesh", mesh_file)
					dest_file = os.path.join(dest_dir, mesh_file)
					monster_tasks.append(copy_file_async(src_file, dest_file, semaphore))
					files_copied = True
			
			if not files_copied:
				shutil.rmtree(dest_dir)
				log_error(f"Removed empty directory: {dest_dir}")
		
		await asyncio.gather(*monster_tasks)
	
	if ani_sheet in workbook.sheetnames:
		ani_ws = workbook[ani_sheet]
		logger.debug(colored(f"Found sheet: {ani_sheet}", 'cyan'))
		
		ani_tasks = []
		for row in ani_ws.iter_rows(min_row = 2, values_only = True):
			monster_code = row[1]
			if not monster_code:
				log_error(f"Missing Monster_Code in row: {row}")
				continue
			dest_dir = os.path.join(r"B:\\Kathana-Out\Sorted", version_name, "Monster", monster_code)
			ensure_directory_exists(dest_dir)
			
			files_copied = False
			for header, ani_file in zip(ani_ws[1][2:], row[2:]):
				if ani_file:
					src_file = os.path.join(version_path, "resource", "object", "Monster", "Ani", ani_file)
					dest_file = os.path.join(dest_dir, ani_file)
					ani_tasks.append(copy_file_async(src_file, dest_file, semaphore))
					files_copied = True
			
			if not files_copied:
				shutil.rmtree(dest_dir)
				log_error(f"Removed empty directory: {dest_dir}")
		
		await asyncio.gather(*ani_tasks)


def generate_fbx_files(
		version_path, entity_type, generate_batch_only = False, combined_batch = False, batch_commands = []
		):
	logger.debug(
			colored(
					f"Entering generate_fbx_files function with version_path: {version_path} and entity_type: {entity_type}",
					'cyan'
					)
			)
	logger.info(f"Generating {entity_type} FBX files from {version_path}...")
	
	root_dir = os.path.join(r"B:\\Kathana-Out\Sorted", os.path.basename(version_path), entity_type)
	fbx_base_dir = os.path.join(r"B:\\Kathana-Out\FBX", os.path.basename(version_path), entity_type)
	ensure_directory_exists(fbx_base_dir)
	
	if combined_batch:
		for root, dirs, files in os.walk(root_dir):
			for file in files:
				if file.endswith(".tmb"):
					tmb_path = os.path.join(root, file)
					tab_files = [f for f in files if f.endswith(".tab")]
					for tab_file in tab_files:
						tab_path = os.path.join(root, tab_file)
						output_file = os.path.join(fbx_base_dir, os.path.relpath(tab_path, root_dir)).replace(
								".tab", ".fbx"
								)
						ensure_directory_exists(os.path.dirname(output_file))
						command = f'"{NOESIS_EXE_PATH}" ?cmode "{tmb_path}" "{output_file}" -loadanimsingle "{tab_path}" -export -bakeanimscale -showstats -animbonenamematch -fbxnoextraframe'
						batch_commands.append(command)
	else:
		batch_file_path = os.path.join(root_dir, f"generate_{entity_type.lower()}_fbx.bat")
		
		with open(batch_file_path, 'w') as batch_file:
			for root, dirs, files in os.walk(root_dir):
				for file in files:
					if file.endswith(".tmb"):
						tmb_path = os.path.join(root, file)
						tab_files = [f for f in files if f.endswith(".tab")]
						for tab_file in tab_files:
							tab_path = os.path.join(root, tab_file)
							output_file = os.path.join(fbx_base_dir, os.path.relpath(tab_path, root_dir)).replace(
									".tab", ".fbx"
									)
							ensure_directory_exists(os.path.dirname(output_file))
							command = f'"{NOESIS_EXE_PATH}" ?cmode "{tmb_path}" "{output_file}" -loadanimsingle "{tab_path}" -export -bakeanimscale -showstats -animbonenamematch -fbxnoextraframe'
							batch_file.write(command + '\n')
		
		logger.info(f"Batch script for generating {entity_type} FBX files created at {batch_file_path}")
		if not generate_batch_only:
			os.system(f'cmd /c "{batch_file_path}"')
			logger.info(f"{entity_type} FBX files generation complete.")


def generate_combined_fbx_batch_file(version_path):
	logger.debug(
			colored(f"Entering generate_combined_fbx_batch_file function with version_path: {version_path}", 'cyan')
			)
	batch_commands = []
	generate_fbx_files(
			version_path, 'PC', generate_batch_only = True, combined_batch = True, batch_commands = batch_commands
			)
	generate_fbx_files(
			version_path, 'NPC', generate_batch_only = True, combined_batch = True, batch_commands = batch_commands
			)
	generate_fbx_files(
			version_path, 'Monster', generate_batch_only = True, combined_batch = True, batch_commands = batch_commands
			)
	
	combined_batch_file_path = os.path.join(
			r"B:\\Kathana-Out\Sorted", os.path.basename(version_path), "generate_all_fbx.bat"
			)
	
	with open(combined_batch_file_path, 'w') as batch_file:
		for command in batch_commands:
			batch_file.write(command + '\n')
	
	logger.info(f"Combined batch script for generating all entity FBX files created at {combined_batch_file_path}")


def clean_up():
	logger.debug(colored(f"Entering clean_up function", 'cyan'))
	sorted_path = r"B:\\Kathana-Out\\Sorted"
	fbx_path = r"B:\\Kathana-Out\\FBX"
	if os.path.exists(sorted_path):
		shutil.rmtree(sorted_path)
		logger.info(colored("Cleaned up the kathana-res-sorted folder", "green"))
	else:
		logger.info(colored("kathana-res-sorted folder does not exist", "yellow"))
	if os.path.exists(fbx_path):
		shutil.rmtree(fbx_path)
		logger.info(colored("Cleaned up the kathana-res-fbx folder", "green"))
	else:
		logger.info(colored("kathana-res-fbx folder does not exist", "yellow"))


def display_menu(chosen_version):
	os.system('cls' if os.name == 'nt' else 'clear')
	logger.info(colored(BANNER, "yellow"))
	if chosen_version:
		version_name = os.path.basename(chosen_version)
		logger.info(colored("_________________________________", "cyan"))
		logger.info(colored(f"Version Selected: {version_name}", "cyan"))
		logger.info(colored("_________________________________", "cyan"))
	else:
		logger.info(colored("No version selected", "red"))
	logger.info(colored("Menu Options:", "blue"))
	logger.info("1 - Choose Kathana Version")
	logger.info("2 - Copy and Sort PC Files")
	logger.info("3 - Copy and Sort NPC Files")
	logger.info("4 - Copy and Sort Monster Files")
	logger.info("5 - Copy and Sort All Entity Files")
	logger.info("6 - Generate PC FBX Files")
	logger.info("6B - Generate PC FBX Batch File Only")
	logger.info("7 - Generate NPC FBX Files")
	logger.info("7B - Generate NPC FBX Batch File Only")
	logger.info("8 - Generate Monster FBX Files")
	logger.info("8B - Generate Monster FBX Batch File Only")
	logger.info("9 - Generate All Entity FBX Files")
	logger.info("9B - Generate All Entity FBX Batch Files Only")
	logger.info("C - Clean Up")
	logger.info("X - Exit")


def choose_version():
	logger.debug("Entering choose_version function")
	logger.info(colored("Choose Kathana Version:", "blue"))
	for idx, version in enumerate(KATHANA_VERSIONS, start = 1):
		logger.info(f"{idx} - {version}")
	choice = input("Enter your choice: ")
	chosen_version = KATHANA_VERSIONS[int(choice) - 1] if choice.isdigit() and 1 <= int(choice) <= len(
			KATHANA_VERSIONS
			) else None
	logger.debug(f"Chosen version: {chosen_version}")
	return chosen_version


def copy_and_sort_files(version_path, entity_type):
	logger.debug(
			f"Entering copy_and_sort_files function with version_path: {version_path} and entity_type: {entity_type}"
			)
	logger.info(f"Copying and sorting {entity_type} files from {version_path}...")
	wb = openpyxl.load_workbook(ENTITY_XLSX_PATH)
	
	start_time = time.time()
	
	with ThreadPoolExecutor(max_workers = 100) as executor:
		futures = []
		if entity_type == 'PC':
			futures.append(executor.submit(asyncio.run, copy_pc_files(wb, version_path)))
		elif entity_type == 'NPC':
			futures.append(executor.submit(asyncio.run, copy_npc_files(wb, version_path)))
		elif entity_type == 'Monster':
			futures.append(executor.submit(asyncio.run, copy_monster_files(wb, version_path)))
		elif entity_type == 'All':
			futures.append(executor.submit(asyncio.run, copy_pc_files(wb, version_path)))
			futures.append(executor.submit(asyncio.run, copy_npc_files(wb, version_path)))
			futures.append(executor.submit(asyncio.run, copy_monster_files(wb, version_path)))
		
		for future in as_completed(futures):
			future.result()
	
	end_time = time.time()
	elapsed_time = end_time - start_time
	logger.info(f"{entity_type} files copied and sorted. Time elapsed: {str(timedelta(seconds = elapsed_time))}")


def main():
	chosen_version = None
	
	while True:
		display_menu(chosen_version)
		choice = input("Enter your choice: ").upper()
		
		if choice == '1':
			chosen_version = choose_version()
			if chosen_version:
				logger.info(colored(f"Chosen version: {chosen_version}", 'cyan'))
			else:
				logger.error("Invalid choice. Please try again.")
		elif choice == '2':
			if chosen_version:
				copy_and_sort_files(chosen_version, 'PC')
			else:
				logger.error("Please choose a Kathana version first.")
		elif choice == '3':
			if chosen_version:
				copy_and_sort_files(chosen_version, 'NPC')
			else:
				logger.error("Please choose a Kathana version first.")
		elif choice == '4':
			if chosen_version:
				copy_and_sort_files(chosen_version, 'Monster')
			else:
				logger.error("Please choose a Kathana version first.")
		elif choice == '5':
			if chosen_version:
				copy_and_sort_files(chosen_version, 'PC')
				copy_and_sort_files(chosen_version, 'NPC')
				copy_and_sort_files(chosen_version, 'Monster')
			else:
				logger.error("Please choose a Kathana version first.")
		elif choice == '6':
			if chosen_version:
				generate_fbx_files(chosen_version, 'PC')
			else:
				logger.error("Please choose a Kathana version first.")
		elif choice == '6B':
			if chosen_version:
				generate_fbx_files(chosen_version, 'PC', generate_batch_only = True)
			else:
				logger.error("Please choose a Kathana version first.")
		elif choice == '7':
			if chosen_version:
				generate_fbx_files(chosen_version, 'NPC')
			else:
				logger.error("Please choose a Kathana version first.")
		elif choice == '7B':
			if chosen_version:
				generate_fbx_files(chosen_version, 'NPC', generate_batch_only = True)
			else:
				logger.error("Please choose a Kathana version first.")
		elif choice == '8':
			if chosen_version:
				generate_fbx_files(chosen_version, 'Monster')
			else:
				logger.error("Please choose a Kathana version first.")
		elif choice == '8B':
			if chosen_version:
				generate_fbx_files(chosen_version, 'Monster', generate_batch_only = True)
			else:
				logger.error("Please choose a Kathana version first.")
		elif choice == '9':
			if chosen_version:
				generate_fbx_files(chosen_version, 'PC')
				generate_fbx_files(chosen_version, 'NPC')
				generate_fbx_files(chosen_version, 'Monster')
			else:
				logger.error("Please choose a Kathana version first.")
		elif choice == '9B':
			if chosen_version:
				generate_combined_fbx_batch_file(chosen_version)
			else:
				logger.error("Please choose a Kathana version first.")
		elif choice == 'C':
			clean_up()
		elif choice == 'X':
			logger.info("Exiting the program.")
			break
		else:
			logger.error("Invalid choice. Please try again.")


if __name__ == "__main__":
	main()
