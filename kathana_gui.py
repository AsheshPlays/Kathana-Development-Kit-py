import sys
import os
from PyQt5.QtWidgets import QApplication, QWidget, QVBoxLayout, QHBoxLayout, QPushButton, QLabel, QMessageBox,\
	QFileDialog, QSizePolicy, QTextEdit, QSpacerItem
from PyQt5.QtGui import QPixmap, QIcon, QPalette, QColor, QFont, QPainter, QPolygon
from PyQt5.QtCore import Qt, QProcess, QThread, pyqtSignal, QPoint
from openpyxl import Workbook
import shutil

# Define display names for the Kathana versions
KATHANA_DISPLAY_NAMES = ["Kathana Global", "Kathana 2", "Kathana 3", "Kathana 3.2", "Kathana 4", "Kathana 5.2",
                         "Kathana 6"]
# Map the display names to the actual paths
KATHANA_VERSIONS = [r"B:\\Kathana\\Kathana-Global", r"B:\\Kathana\\Kathana2", r"B:\\Kathana\\Kathana3",
                    r"B\\Kathana\\Kathana3.2", r"B\\Kathana\\Kathana4", r"B\\Kathana\\Kathana5.2",
                    r"B\\Kathana\\Kathana6"]

LOG_XLSX_FILENAME = "KATHANA_LOGS.xlsx"  # Set the log filename without timestamp
LOG_XLSX_PATH = os.path.join(os.getcwd(), LOG_XLSX_FILENAME)


def initialize_log_workbook():
	"""
	Initialize the log workbook with sheets for error and success logs.
	"""
	wb_log = Workbook()
	error_log_ws = wb_log.create_sheet('ERROR_LOGS')
	success_log_ws = wb_log.create_sheet('SUCCESS_LOGS')
	default_sheet = wb_log.active
	wb_log.remove(default_sheet)
	wb_log.save(LOG_XLSX_PATH)
	return wb_log


def ensure_directory_exists(path):
	"""
	Ensure that the specified directory exists, creating it if necessary.
	"""
	if not os.path.exists(path):
		os.makedirs(path)


# Ensure the log directory exists
ensure_directory_exists(os.path.dirname(LOG_XLSX_PATH))

wb_log = initialize_log_workbook()
error_log_ws = wb_log['ERROR_LOGS']
success_log_ws = wb_log['SUCCESS_LOGS']


class Worker(QThread):
	"""
	Worker thread to handle tasks in the background.
	"""
	progress = pyqtSignal(int)
	output = pyqtSignal(str)
	error = pyqtSignal(str)
	finished = pyqtSignal()
	
	def __init__(self, task, *args, **kwargs):
		super().__init__()
		self.task = task
		self.args = args
		self.kwargs = kwargs
	
	def run(self):
		try:
			self.task(*self.args, **self.kwargs)
		except Exception as e:
			self.error.emit(str(e))
		self.finished.emit()


class ArrowButton(QPushButton):
	"""
	Custom button to display an arrow pointing left or right.
	"""
	
	def __init__(self, direction, parent = None):
		super().__init__(parent)
		self.direction = direction
		self.setFixedSize(30, 30)  # Adjust the size as needed
		self.setStyleSheet("background-color: transparent; border: none;")
	
	def paintEvent(self, event):
		painter = QPainter(self)
		painter.setRenderHint(QPainter.Antialiasing)
		painter.setBrush(QColor("red"))
		painter.setPen(Qt.NoPen)
		
		if self.direction == "left":
			points = [QPoint(self.width(), 0), QPoint(0, self.height() // 2), QPoint(self.width(), self.height())]
		elif self.direction == "right":
			points = [QPoint(0, 0), QPoint(self.width(), self.height() // 2), QPoint(0, self.height())]
		
		triangle = QPolygon(points)
		painter.drawPolygon(triangle)


class KathanaVersionTool(QWidget):
	"""
	Main application window for the Kathana Version Tool.
	"""
	
	def __init__(self):
		super().__init__()
		self.selected_index = 0
		self.selected_version = KATHANA_VERSIONS[self.selected_index]
		self.process = QProcess(self)
		self.version_set = False
		self.initUI()
	
	def initUI(self):
		"""
		Initialize the user interface.
		"""
		self.setWindowTitle('Kathana Version Selector')
		self.setFixedSize(800, 600)
		self.setWindowIcon(QIcon('asheshicon.png'))  # Set the application icon
		
		# Set global font
		QApplication.setFont(QFont("Dotum", 8))
		
		layout = QVBoxLayout()
		
		self.banner_label = QLabel(self)
		pixmap = QPixmap('asheshdevkitbanner.png')  # Load the banner image
		self.banner_label.setPixmap(pixmap)
		self.banner_label.setScaledContents(True)  # Ensure the image scales to fill the label
		self.banner_label.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)  # Allow the label to expand
		layout.addWidget(self.banner_label)
		
		# Comment for banner ratio and size
		# The banner is expected to maintain a ratio of width:height = 4:1 for optimal display
		# Example dimensions: 800x200 pixels
		
		version_selector_layout = QHBoxLayout()
		version_selector_layout.addSpacerItem(QSpacerItem(40, 20, QSizePolicy.Expanding, QSizePolicy.Minimum))
		
		left_arrow_btn = ArrowButton("left", self)
		left_arrow_btn.clicked.connect(self.select_previous_version)
		version_selector_layout.addWidget(left_arrow_btn, alignment = Qt.AlignLeft)
		
		self.selected_version_label = QLabel(KATHANA_DISPLAY_NAMES[self.selected_index])
		self.selected_version_label.setAlignment(Qt.AlignCenter)
		self.selected_version_label.setFont(QFont("Dotum", 16, QFont.Bold))
		palette = self.selected_version_label.palette()
		palette.setColor(QPalette.WindowText, Qt.red if self.version_set else QColor("orange"))
		self.selected_version_label.setPalette(palette)
		version_selector_layout.addWidget(self.selected_version_label, alignment = Qt.AlignCenter)
		
		right_arrow_btn = ArrowButton("right", self)
		right_arrow_btn.clicked.connect(self.select_next_version)
		version_selector_layout.addWidget(right_arrow_btn, alignment = Qt.AlignRight)
		
		version_selector_layout.addSpacerItem(QSpacerItem(40, 20, QSizePolicy.Expanding, QSizePolicy.Minimum))
		layout.addLayout(version_selector_layout)
		
		set_version_btn = QPushButton('Set Version', self)
		set_version_btn.clicked.connect(self.set_version)
		layout.addWidget(set_version_btn)
		
		browse_btn = QPushButton('Browse for Kathana Version Directory', self)
		browse_btn.clicked.connect(self.browse_version)
		layout.addWidget(browse_btn)
		
		buttons_layout = QHBoxLayout()
		
		col1_layout = QVBoxLayout()
		self.add_button_row(
			col1_layout, [('Copy and Sort PC Files', lambda: self.run_task('PC')),
						('Copy and Sort NPC Files', lambda: self.run_task('NPC')),
						('Copy and Sort Monster Files', lambda: self.run_task('Monster')),
						('Copy and Sort All Entity Files', lambda: self.run_task('All'))]
			)
		buttons_layout.addLayout(col1_layout)
		
		col2_layout = QVBoxLayout()
		self.add_button_row(
			col2_layout, [('Generate PC FBX Files', lambda: self.run_fbx_task('PC')),
						('Generate NPC FBX Files', lambda: self.run_fbx_task('NPC')),
						('Generate Monster FBX Files', lambda: self.run_fbx_task('Monster')),
						('Generate All Entity FBX Files', lambda: self.run_fbx_task('All'))]
			)
		buttons_layout.addLayout(col2_layout)
		
		col3_layout = QVBoxLayout()
		self.add_button_row(
			col3_layout, [('Generate PC FBX Batch File Only', lambda: self.run_fbx_task('PC', True)),
						('Generate NPC FBX Batch File Only', lambda: self.run_fbx_task('NPC', True)),
						('Generate Monster FBX Batch File Only', lambda: self.run_fbx_task('Monster', True)),
						('Generate All FBX Batch File Only', self.run_combined_fbx_batch_file)]
			)
		buttons_layout.addLayout(col3_layout)
		
		layout.addLayout(buttons_layout)
		
		control_buttons_layout = QHBoxLayout()
		self.add_button_row(
			control_buttons_layout,
			[('Clean Up', self.clean_up), ('Stop', self.stop_processes), ('Restart', self.restart_application)]
			)
		layout.addLayout(control_buttons_layout)
		
		console_section_layout = QVBoxLayout()
		
		console_label = QLabel('Console')
		console_label.setFont(QFont("Dotum", 10, QFont.Bold))
		console_section_layout.addWidget(console_label)
		
		console_layout = QHBoxLayout()
		
		self.console_output = QTextEdit()
		self.console_output.setReadOnly(True)
		console_palette = self.console_output.palette()
		console_palette.setColor(QPalette.Base, QColor(200, 200, 200))  # Set the background color to a darker gray
		self.console_output.setPalette(console_palette)
		console_layout.addWidget(self.console_output)
		
		clear_console_btn = QPushButton('Clear \nConsole')
		clear_console_btn.setStyleSheet("background-color: red; color: white;")
		clear_console_btn.setFont(QFont("Dotum", 10))
		clear_console_btn.clicked.connect(self.clear_console)
		clear_console_btn.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
		clear_console_btn.setFixedWidth(150)  # Adjust the width as needed to make it square
		console_layout.addWidget(clear_console_btn)
		
		console_section_layout.addLayout(console_layout)
		
		layout.addLayout(console_section_layout)
		
		author_version_layout = QHBoxLayout()
		author_label = QLabel('Ashesh Development © 2024')
		author_label.setFont(QFont("Dotum", 8))
		version_label = QLabel('Version: 1.0.4')
		version_label.setFont(QFont("Dotum", 8))
		author_version_layout.addWidget(author_label)
		author_version_layout.addSpacerItem(QSpacerItem(40, 20, QSizePolicy.Expanding, QSizePolicy.Minimum))
		author_version_layout.addWidget(version_label)
		layout.addLayout(author_version_layout)
		
		self.setLayout(layout)
		
		self.process.setProcessChannelMode(QProcess.MergedChannels)
		self.process.readyReadStandardOutput.connect(self.on_readyReadStandardOutput)
		self.process.readyReadStandardError.connect(self.on_readyReadStandardError)
		self.process.finished.connect(self.on_task_finished)
	
	def add_button_row(self, parent_layout, buttons):
		"""
		Add a row of buttons to the specified layout.
		"""
		for label, callback in buttons:
			button = QPushButton(label)
			button.clicked.connect(callback)
			button.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)
			parent_layout.addWidget(button)
	
	def update_selected_version(self):
		"""
		Update the selected Kathana version.
		"""
		self.selected_version = KATHANA_VERSIONS[self.selected_index]
		self.selected_version_label.setText(KATHANA_DISPLAY_NAMES[self.selected_index])
		self.update_version_label_color()
	
	def select_previous_version(self):
		"""
		Select the previous Kathana version in the list.
		"""
		self.selected_index = (self.selected_index - 1) % len(KATHANA_VERSIONS)
		self.update_selected_version()
	
	def select_next_version(self):
		"""
		Select the next Kathana version in the list.
		"""
		self.selected_index = (self.selected_index + 1) % len(KATHANA_VERSIONS)
		self.update_selected_version()
	
	def set_version(self):
		"""
		Set the selected Kathana version and update the label color.
		"""
		self.version_set = True
		self.update_version_label_color()
		QMessageBox.information(
			self, 'Version Set', f'Selected version set to: {KATHANA_DISPLAY_NAMES[self.selected_index]}'
			)
	
	def update_version_label_color(self):
		"""
		Update the color of the version label based on whether the version is set.
		"""
		palette = self.selected_version_label.palette()
		palette.setColor(QPalette.WindowText, Qt.red if self.version_set else QColor("orange"))
		self.selected_version_label.setPalette(palette)
	
	def browse_version(self):
		"""
		Open a file dialog to select a Kathana version directory.
		"""
		options = QFileDialog.Options()
		options |= QFileDialog.ShowDirsOnly
		directory = QFileDialog.getExistingDirectory(self, "Select Kathana Version Directory", options = options)
		if directory:
			self.selected_version = directory
			self.selected_version_label.setText(os.path.basename(directory))
			self.version_set = True
			self.update_version_label_color()
	
	def run_task(self, entity_type):
		"""
		Run a task to copy and sort files for the specified entity type.
		"""
		if self.selected_version:
			self.worker = Worker(self.process_task, entity_type)
			self.worker.output.connect(self.append_output)
			self.worker.error.connect(self.append_error)
			self.worker.finished.connect(self.on_task_finished)
			self.worker.start()
		else:
			QMessageBox.warning(self, 'Error', 'Please choose a Kathana version first.')
	
	def run_fbx_task(self, entity_type, generate_batch_only = False):
		"""
		Run a task to generate FBX files for the specified entity type.
		"""
		if self.selected_version:
			self.worker = Worker(self.process_fbx_task, entity_type, generate_batch_only)
			self.worker.output.connect(self.append_output)
			self.worker.error.connect(self.append_error)
			self.worker.finished.connect(self.on_task_finished)
			self.worker.start()
		else:
			QMessageBox.warning(self, 'Error', 'Please choose a Kathana version first.')
	
	def run_combined_fbx_batch_file(self):
		"""
		Run a task to generate combined FBX batch files.
		"""
		if self.selected_version:
			self.worker = Worker(self.process_combined_fbx_task)
			self.worker.output.connect(self.append_output)
			self.worker.error.connect(self.append_error)
			self.worker.finished.connect(self.on_task_finished)
			self.worker.start()
		else:
			QMessageBox.warning(self, 'Error', 'Please choose a Kathana version first.')
	
	def clean_up(self):
		"""
		Run a task to clean up generated files.
		"""
		self.worker = Worker(self.process_clean_up)
		self.worker.output.connect(self.append_output)
		self.worker.error.connect(self.append_error)
		self.worker.finished.connect(self.on_task_finished)
		self.worker.start()
	
	def stop_processes(self):
		"""
		Stop all running processes.
		"""
		self.process.kill()
		QMessageBox.information(self, 'Process Stopped', 'All running processes have been stopped.')
	
	def restart_application(self):
		"""
		Restart the application.
		"""
		QApplication.quit()
		os.execl(sys.executable, sys.executable, *sys.argv)
	
	def process_task(self, entity_type):
		"""
		Process task to copy and sort files.
		"""
		ensure_directory_exists(self.selected_version)
		self.process.setWorkingDirectory(os.getcwd())
		self.process.start(
				sys.executable, ['-c',
				                 f'import kathana_clx; kathana_clx.copy_and_sort_files("{self.selected_version}", "{entity_type}")']
				)
		self.process.waitForFinished()
	
	def process_fbx_task(self, entity_type, generate_batch_only):
		"""
		Process task to generate FBX files.
		"""
		ensure_directory_exists(self.selected_version)
		self.process.setWorkingDirectory(os.getcwd())
		self.process.start(
				sys.executable, ['-c',
				                 f'import kathana_clx; kathana_clx.generate_fbx_files("{self.selected_version}", "{entity_type}", {generate_batch_only})']
				)
		self.process.waitForFinished()
	
	def process_combined_fbx_task(self):
		"""
		Process task to generate combined FBX batch files.
		"""
		ensure_directory_exists(self.selected_version)
		self.process.setWorkingDirectory(os.getcwd())
		self.process.start(
				sys.executable,
				['-c', f'import kathana_clx; kathana_clx.generate_combined_fbx_batch_file("{self.selected_version}")']
				)
		self.process.waitForFinished()
	
	def process_clean_up(self):
		"""
		Process task to clean up files.
		"""
		try:
			sorted_path = r"B:\\Kathana-Out\\Sorted"
			fbx_path = r"B:\\Kathana-Out\\FBX"
			if os.path.exists(sorted_path):
				shutil.rmtree(sorted_path)
				self.append_output("Cleaned up the kathana-res-sorted folder")
			else:
				self.append_output("kathana-res-sorted folder does not exist")
			if os.path.exists(fbx_path):
				shutil.rmtree(fbx_path)
				self.append_output("Cleaned up the kathana-res-fbx folder")
			else:
				self.append_output("kathana-res-fbx folder does not exist")
		except Exception as e:
			self.append_error(f"Error during clean up: {str(e)}")
	
	def clear_console(self):
		"""
		Clear the console output.
		"""
		self.console_output.clear()
	
	def append_output(self, text):
		"""
		Append output text to the console.
		"""
		self.console_output.setTextColor(QColor("blue"))
		self.console_output.append(text)
		self.console_output.setTextColor(QColor("black"))
	
	def append_error(self, text):
		"""
		Append error text to the console.
		"""
		self.console_output.setTextColor(QColor("red"))
		self.console_output.append(f"ERROR: {text}")
		self.console_output.setTextColor(QColor("black"))
	
	def on_task_finished(self):
		"""
		Handle task finished event.
		"""
		self.console_output.setTextColor(QColor("green"))
		self.console_output.append("Task Finished: The task has been completed successfully.")
		self.console_output.setTextColor(QColor("black"))
	
	def on_readyReadStandardOutput(self):
		"""
		Handle ready read standard output event.
		"""
		text = self.process.readAllStandardOutput().data().decode()
		self.append_output(text)
	
	def on_readyReadStandardError(self):
		"""
		Handle ready read standard error event.
		"""
		text = self.process.readAllStandardError().data().decode()
		self.append_error(text)


if __name__ == '__main__':
	app = QApplication(sys.argv)
	tool = KathanaVersionTool()
	tool.show()
	sys.exit(app.exec_())
